import os as custom_os
import time
import difflib
import openpyxl
import ast

print2log("===== NODE 5: Write to Excel (FIXED v2) =====")


def normalize_header(text):
    """Aggressively normalize a header string for matching: lowercase,
    collapse whitespace, strip common punctuation/degree-symbol variants,
    and unify British/American spelling differences we've seen in these
    templates (e.g. 'Magnetising' vs 'Magnetizing')."""
    s = str(text).strip().lower()
    s = s.replace("°", "").replace("º", "")          # degree symbol variants
    s = s.replace("magnetising", "magnetizing")       # spelling variant
    s = s.replace("bi-dir.", "bidirectional").replace("bi-dir", "bidirectional")
    s = s.replace("_", " ").replace("-", " ")
    s = " ".join(s.split())                            # collapse whitespace
    # strip trailing punctuation like ':' '.' that headers sometimes have
    s = s.strip(" :.")
    return s


def write_data_to_excel(xlsx_path, data_dict, sheet_name_hint, header_row=2, data_row=4,
                         max_retries=5, retry_delay_sec=3):
    if not xlsx_path or not custom_os.path.isfile(xlsx_path):
        print2log(f"ERROR: Excel path not found, skipping write: {xlsx_path}")
        return False

    basename = custom_os.path.basename(xlsx_path)
    if basename.startswith("~$"):
        real_name = basename[2:]
        real_path = custom_os.path.join(custom_os.path.dirname(xlsx_path), real_name)
        if custom_os.path.isfile(real_path):
            print2log(f"WARNING: '{basename}' is an Office lock/temp file. Falling back to "
                       f"actual file: '{real_name}'")
            xlsx_path = real_path
        else:
            print2log(f"ERROR: '{xlsx_path}' is an Office lock/temp file with no real workbook "
                       f"alongside it. Close it in Excel (or delete the stray lock file) and re-run.")
            return False

    if not data_dict:
        print2log(f"WARNING: No data to write for '{xlsx_path}' — skipping write.")
        return False

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        print2log(f"ERROR: Could not open workbook '{xlsx_path}': {e}")
        return False

    ws = wb.worksheets[0]
    print2log(f"Opened sheet '{ws.title}' in '{custom_os.path.basename(xlsx_path)}'")

    # ── Build header map using BOTH an exact-normalized key and a raw list
    # of (normalized_header, column_index) for fuzzy fallback matching ─────
    header_map = {}
    header_list = []  # [(normalized_header, col_idx, original_text)]
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=header_row, column=col).value
        if cell_val:
            norm = normalize_header(cell_val)
            header_map[norm] = col
            header_list.append((norm, col, str(cell_val).strip()))

    written = 0
    not_matched = []
    for header_label, value in data_dict.items():
        key = normalize_header(header_label)
        col_idx = header_map.get(key)

        # ── Fallback 1: fuzzy closest-match against all real headers ───────
        if not col_idx and header_list:
            candidates = [h[0] for h in header_list]
            close = difflib.get_close_matches(key, candidates, n=1, cutoff=0.82)
            if close:
                match_norm = close[0]
                for norm, idx, orig in header_list:
                    if norm == match_norm:
                        col_idx = idx
                        print2log(f"  [FUZZY MATCH] '{header_label}' -> Excel column '{orig}' "
                                   f"(similarity match, not exact)")
                        break

        if col_idx:
            ws.cell(row=data_row, column=col_idx, value=value)
            written += 1
        else:
            not_matched.append(header_label)

    if not_matched:
        print2log(f"  WARNING: {len(not_matched)} field(s) had no matching Excel column: {not_matched}")
        print2log(f"  Actual Excel headers in row {header_row}: {[h[2] for h in header_list]}")

    # ── Retry-with-backoff on transient OS-level file locks ────────────────
    # "[Errno 16] Device or resource busy" means the OS/filesystem itself
    # has an exclusive lock on the file at save time (e.g. an antivirus
    # scan, a sync client, or another process briefly touching the file on
    # a network share). This is NOT something openpyxl or this script can
    # force past directly — but if the lock is transient, a short retry
    # loop resolves it without failing the whole pipeline run. If the file
    # is genuinely open in Excel by someone, every retry will still fail
    # and the real error will be logged after the final attempt.
    last_error = None
    for attempt in range(1, max_retries + 1):
        try:
            wb.save(xlsx_path)
            print2log(f"SUCCESS: Wrote {written} field(s) to '{custom_os.path.basename(xlsx_path)}' "
                       f"({len(not_matched)} unmatched). Saved at row {data_row}.")
            return True
        except OSError as e:
            last_error = e
            if attempt < max_retries:
                print2log(f"  WARNING: Save attempt {attempt}/{max_retries} failed "
                           f"(file busy/locked): {e}. Retrying in {retry_delay_sec}s...")
                time.sleep(retry_delay_sec)
            else:
                print2log(f"ERROR: Could not save workbook '{xlsx_path}' after {max_retries} "
                           f"attempts: {e}. The file is likely open in Excel or locked by another "
                           f"process/sync tool on the storage share — this needs to be closed by "
                           f"a person before the pipeline can write to it.")
        except Exception as e:
            print2log(f"ERROR: Could not save workbook '{xlsx_path}': {e}")
            return False

    return False


# ── Rubiscape entry point ─────────────────────────────────────────────────────
print2log(f"inputData keys: {list(inputData.keys())}")

df = inputData.get("Parse_Pump_Data_PDS")
if df is None or len(df) == 0:
    print2log("ERROR: No data received from Node 4 (Parse_Pump_Data_PDS)")
    raise Exception("NODE 5: missing input DataFrame from predecessor task")

row = df.iloc[0]

motor_xlsx_path = row.get("motor_xlsx_path")
pump_xlsx_path  = row.get("pump_xlsx_path")
motor_data       = row.get("motor_data")
pump_data        = row.get("pump_data")

if isinstance(motor_data, str):
    try:
        motor_data = ast.literal_eval(motor_data)
    except Exception:
        print2log("WARNING: Could not parse motor_data string back into dict")
        motor_data = {}

if isinstance(pump_data, str):
    try:
        pump_data = ast.literal_eval(pump_data)
    except Exception:
        print2log("WARNING: Could not parse pump_data string back into dict")
        pump_data = {}

motor_status = write_data_to_excel(motor_xlsx_path, motor_data, "motor")
pump_status  = write_data_to_excel(pump_xlsx_path, pump_data, "pump")

print2log(f"Motor excel write status: {motor_status}")
print2log(f"Pump excel write status: {pump_status}")
print2log("===== NODE 5 COMPLETE — PIPELINE FINISHED =====")

output_df = df.copy()
output_df["motor_write_status"] = motor_status
output_df["pump_write_status"]  = pump_status

return output_df
